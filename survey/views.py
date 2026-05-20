"""
survey/views.py
"""
from django.shortcuts import render, redirect
from django.http import Http404
from .services.data_loader import scan_routes, load_route_data, build_tour, get_alignment_coords, get_route_length_m


def overview(request):
    routes = scan_routes()
    route_cards = []

    for route_id, info in routes.items():
        try:
            df_features, df_pp, summary = load_route_data(info["xlsx"])
            valid  = df_features.dropna(subset=["Latitude", "Longitude"])
            coords = get_alignment_coords(info, df_features)
            route_cards.append({
                "route_id":   route_id,
                "info":       info,
                "n_features": len(df_features),
                "n_pp":       len(df_pp),
                "n_good":     int((df_features["Condition"] == "GOOD").sum()),
                "n_fair":     int((df_features["Condition"] == "FAIR").sum()),
                "n_poor":     int((df_features["Condition"] == "POOR").sum()),
                "ch_min":     float(df_features["Chainage (m)"].min()),
                "ch_max":     float(df_features["Chainage (m)"].max()),
                "has_dxf":    bool(info["dxf"]),
                "center_lat": float(valid["Latitude"].mean()) if not valid.empty else None,
                "center_lon": float(valid["Longitude"].mean()) if not valid.empty else None,
                "coords":     coords,
            })
        except Exception as exc:
            route_cards.append({"route_id": route_id, "info": info, "error": str(exc)})

    import json

    all_lats = [c["center_lat"] for c in route_cards if c.get("center_lat")]
    all_lons = [c["center_lon"] for c in route_cards if c.get("center_lon")]

    # Build per-route bounds and JSON data for Leaflet map
    map_routes = []
    for card in route_cards:
        if card.get("error") or not card.get("coords"):
            continue
        lats = [c[0] for c in card["coords"]]
        lons = [c[1] for c in card["coords"]]
        bounds = {
            "lat_min": min(lats) - 0.005,
            "lat_max": max(lats) + 0.005,
            "lon_min": min(lons) - 0.005,
            "lon_max": max(lons) + 0.005,
        }
        card["bounds"] = bounds
        map_routes.append({
            "route_id":   card["route_id"],
            "label":      card["info"]["label"],
            "coords":     card["coords"],
            "center":     [card["center_lat"], card["center_lon"]],
            "bounds":     bounds,
            "n_features": card["n_features"],
            "n_pp":       card["n_pp"],
            "n_good":     card["n_good"],
            "n_fair":     card["n_fair"],
            "n_poor":     card["n_poor"],
        })

    all_bounds = None
    if all_lats:
        all_bounds = {
            "lat_min": min(all_lats) - 0.01,
            "lat_max": max(all_lats) + 0.01,
            "lon_min": min(all_lons) - 0.01,
            "lon_max": max(all_lons) + 0.01,
        }

    first_route = list(routes.keys())[0] if routes else None
    return render(request, "survey/overview.html", {
        "routes":      routes,
        "route_cards": route_cards,
        "map_routes":  json.dumps(map_routes),
        "all_bounds":  json.dumps(all_bounds) if all_bounds else "null",
        "first_route": first_route,
    })


def route_detail(request, route_id):
    import json
    import math

    routes = scan_routes()
    if route_id not in routes:
        raise Http404(f"Route {route_id} not found")

    info     = routes[route_id]
    df_features, df_pp, summary = load_route_data(info["xlsx"])

    def safe(val):
        """Convert numpy/nan values to JSON-safe Python types."""
        if val is None:
            return ""
        try:
            if math.isnan(float(val)):
                return ""
        except (TypeError, ValueError):
            pass
        return val

    # Features as JSON for Leaflet markers + inspect panel
    features_json = []
    for _, row in df_features.iterrows():
        if not (safe(row.get("Latitude")) and safe(row.get("Longitude"))):
            continue
        features_json.append({
            "id":         str(row["ID"]),
            "type":       str(row.get("Feature Type", "")),
            "condition":  str(row.get("Condition", "")),
            "side":       str(row.get("Side", "")),
            "chainage":   safe(row.get("Chainage (m)")),
            "offset":     safe(row.get("Offset from Edge (m)")),
            "easting":    safe(row.get("Easting", "")),
            "northing":   safe(row.get("Northing", "")),
            "gps_acc":    safe(row.get("GPS Accuracy (m)")),
            "notes":      str(row.get("Notes", "") or ""),
            "captured_by": str(row.get("Captured By", "") or ""),
            "captured_at": str(row.get("Captured At", "") or "")[:16],
            "lat":        float(row["Latitude"]),
            "lon":        float(row["Longitude"]),
            "kind":       "feature",
        })

    # Passing places as JSON
    pp_json = []
    for _, row in df_pp.iterrows():
        if not (safe(row.get("Mid Latitude")) and safe(row.get("Mid Longitude"))):
            continue
        pp_json.append({
            "id":         str(row["PP ID"]),
            "type":       "Passing Place",
            "condition":  str(row.get("Status", "")),
            "side":       str(row.get("Side", "")),
            "chainage":   safe(row.get("Mid Chainage (m)")),
            "width":      safe(row.get("Width (m)")),
            "length":     safe(row.get("Length (m)")),
            "easting":    safe(row.get("Mid Easting", "")),
            "northing":   safe(row.get("Mid Northing", "")),
            "gps_acc":    safe(row.get("GPS Accuracy (m)")),
            "notes":      str(row.get("Notes", "") or ""),
            "captured_by": str(row.get("Captured By", "") or ""),
            "captured_at": str(row.get("Captured At", "") or "")[:16],
            "lat":        float(row["Mid Latitude"]),
            "lon":        float(row["Mid Longitude"]),
            "kind":       "pp",
        })

    # Route alignment — DXF if available, GPS fallback
    route_coords = get_alignment_coords(info, df_features)

    # Sort both tables by chainage
    df_features  = df_features.sort_values("Chainage (m)").reset_index(drop=True)
    df_pp        = df_pp.sort_values("Mid Chainage (m)").reset_index(drop=True)

    context = {
        "routes":        routes,
        "first_route":   list(routes.keys())[0] if routes else None,
        "route_id":      route_id,
        "info":          info,
        "features":      df_features.to_dict("records"),
        "passing_places": df_pp.to_dict("records"),
        "features_json": json.dumps(features_json),
        "pp_json":       json.dumps(pp_json),
        "route_coords":  json.dumps(route_coords),
        "n_features":    len(df_features),
        "n_pp":          len(df_pp),
        "n_good":        int((df_features["Condition"] == "GOOD").sum()),
        "n_fair":        int((df_features["Condition"] == "FAIR").sum()),
        "n_poor":        int((df_features["Condition"] == "POOR").sum()),
    }
    return render(request, "survey/route_detail.html", context)


def route_report(request, route_id):
    from .services.map_renderer import CACHE_DIR

    routes = scan_routes()
    if route_id not in routes:
        raise Http404(f"Route {route_id} not found")

    info = routes[route_id]
    df_features, df_pp, summary = load_route_data(info['xlsx'])
    tour = build_tour(df_features, df_pp)

    n_good = int((df_features['Condition'] == 'GOOD').sum())
    n_fair = int((df_features['Condition'] == 'FAIR').sum())
    n_poor = int((df_features['Condition'] == 'POOR').sum())

    # Route alignment — DXF if available, GPS fallback
    route_coords = get_alignment_coords(info, df_features)

    # Attach cached map URLs — no generation here, run: python manage.py generate_maps
    for stop in tour:
        safe_id  = str(stop['id']).replace('/', '_').replace(' ', '_')
        overview = CACHE_DIR / f"{route_id}_{safe_id}_overview.png"
        detail   = CACHE_DIR / f"{route_id}_{safe_id}_detail.png"
        stop['overview_map_url'] = f"/media/map_cache/{route_id}_{safe_id}_overview.png" if overview.exists() else None
        stop['detail_map_url']   = f"/media/map_cache/{route_id}_{safe_id}_detail.png"   if detail.exists()   else None

    maps_ready = any(s['overview_map_url'] for s in tour)

    context = {
        'routes':     routes,
        'first_route': list(routes.keys())[0] if routes else None,
        'route_id':   route_id,
        'info':       info,
        'tour':       tour,
        'n_features': len(df_features),
        'n_pp':       len(df_pp),
        'n_good':     n_good,
        'n_fair':     n_fair,
        'n_poor':     n_poor,
        'ch_min':     tour[0]['chainage'] if tour else 0,
        'ch_max':     tour[-1]['chainage'] if tour else 0,
        'maps_ready': maps_ready,
    }
    return render(request, 'survey/report.html', context)


def route_summary(request, route_id):
    routes = scan_routes()
    if route_id not in routes:
        raise Http404(f"Route {route_id} not found")

    info = routes[route_id]
    df_features, df_pp, summary = load_route_data(info['xlsx'])

    context = {
        'routes':     routes,
        'route_id':   route_id,
        'info':       info,
        'summary':    summary,
        'n_features': len(df_features),
        'n_pp':       len(df_pp),
    }
    return render(request, 'survey/summary.html', context)

def photo_library(request):
    import json
    import io
    import zipfile
    from django.http import HttpResponse
    from .services.data_loader import feature_color

    routes = scan_routes()
    if not routes:
        return render(request, 'survey/photo_library.html', {'routes': routes})

    # Route selection
    route_id = request.GET.get('route') or request.POST.get('route') or list(routes.keys())[0]
    if route_id not in routes:
        route_id = list(routes.keys())[0]

    info = routes[route_id]
    df_features, df_pp, summary = load_route_data(info['xlsx'])

    # ── Filters ───────────────────────────────────────────────
    q_id      = request.GET.get('q_id', '').strip().lower()
    q_type    = request.GET.get('q_type', '')
    ch_min    = request.GET.get('ch_min', '')
    ch_max    = request.GET.get('ch_max', '')
    photos_only = request.GET.get('photos_only', '') == '1'

    def apply_filters(df, id_col, chainage_col, type_col=None):
        if q_id:
            df = df[df[id_col].astype(str).str.lower().str.contains(q_id)]
        if q_type and type_col:
            df = df[df[type_col] == q_type]
        if ch_min:
            try: df = df[df[chainage_col] >= float(ch_min)]
            except: pass
        if ch_max:
            try: df = df[df[chainage_col] <= float(ch_max)]
            except: pass
        return df

    df_f  = apply_filters(df_features.copy(), 'ID', 'Chainage (m)', 'Feature Type')
    df_pp = apply_filters(df_pp.copy(), 'PP ID', 'Mid Chainage (m)')

    # ── Build photo items ─────────────────────────────────────
    import math
    def safe_str(val):
        if val is None: return ''
        try:
            if math.isnan(float(val)): return ''
        except: pass
        return str(val).strip()

    def make_new_name(route_id, item_id, feature_type, condition, chainage, original, index=None):
        ext   = original.rsplit('.', 1)[-1] if '.' in original else 'jpg'
        ch    = f"CH{float(chainage):.0f}m" if chainage else ''
        idx   = f"_{index}" if index is not None else ''
        parts = [p for p in [route_id, item_id, feature_type, condition, ch] if p]
        return '_'.join(parts).replace(' ', '_').replace('/', '_') + idx + '.' + ext

    items = []
    for _, row in df_f.iterrows():
        photos = [p.strip() for p in str(row.get('Photo', '') or '').split(',') if p.strip()]
        if photos_only and not photos:
            continue
        for i, photo in enumerate(photos):
            from pathlib import Path
            from django.conf import settings
            path = settings.MEDIA_ROOT / 'photos' / 'features' / photo
            new_name = make_new_name(
                route_id, safe_str(row['ID']),
                safe_str(row.get('Feature Type', '')),
                safe_str(row.get('Condition', '')),
                row.get('Chainage (m)', ''), photo,
                index=(i + 1) if len(photos) > 1 else None
            )
            items.append({
                'id':           safe_str(row['ID']),
                'kind':         'feature',
                'type':         safe_str(row.get('Feature Type', '')),
                'condition':    safe_str(row.get('Condition', '')),
                'chainage':     row.get('Chainage (m)', ''),
                'side':         safe_str(row.get('Side', '')),
                'photo':        photo,
                'new_name':     new_name,
                'has_photo':    path.exists(),
                'photo_url':    f'/media/photos/features/{photo}' if path.exists() else None,
                'captured_at':  safe_str(row.get('Captured At', ''))[:10],
            })


    for _, row in df_pp.iterrows():
        photos = [p.strip() for p in str(row.get('Photo', '') or '').split(',') if p.strip()]
        if photos_only and not photos:
            continue
        for i, photo in enumerate(photos):
            from pathlib import Path
            from django.conf import settings
            path = settings.MEDIA_ROOT / 'photos' / 'passing_places' / photo
            new_name = make_new_name(
                route_id, safe_str(row['PP ID']),
                'PassingPlace',
                safe_str(row.get('Status', '')),
                row.get('Mid Chainage (m)', ''), photo,
                index=(i + 1) if len(photos) > 1 else None
            )
            items.append({
                'id':           safe_str(row['PP ID']),
                'kind':         'pp',
                'type':         'Passing Place',
                'condition':    safe_str(row.get('Status', '')),
                'chainage':     row.get('Mid Chainage (m)', ''),
                'side':         safe_str(row.get('Side', '')),
                'photo':        photo,
                'new_name':     new_name,
                'has_photo':    path.exists(),
                'photo_url':    f'/media/photos/passing_places/{photo}' if path.exists() else None,
                'captured_at':  safe_str(row.get('Captured At', ''))[:10],
            })


    # Keep only items with photos and sort by chainage
    items = [i for i in items if i['has_photo']]
    items.sort(key=lambda x: float(x['chainage']) if x['chainage'] else 0)

    # ── Download zip of selected photos ──────────────────────
    if request.method == 'POST' and request.POST.get('action') == 'download_photos':
        selected = request.POST.getlist('selected_photos')
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            used_names = {}
            for item in items:
                if item['photo'] in selected and item['has_photo']:
                    subfolder = 'passing_places' if item['kind'] == 'pp' else 'features'
                    src = settings.MEDIA_ROOT / 'photos' / subfolder / item['photo']
                    # Guarantee unique name in zip
                    name = item['new_name']
                    if name in used_names:
                        base, ext = name.rsplit('.', 1)
                        used_names[name] += 1
                        name = f"{base}_{used_names[name]}.{ext}"
                    else:
                        used_names[name] = 1
                    zf.write(src, name)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/zip')
        resp['Content-Disposition'] = f'attachment; filename="{route_id}_photos.zip"'
        return resp

    # ── Download Excel ────────────────────────────────────────
    if request.GET.get('action') == 'download_excel':
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        wb = openpyxl.Workbook()

        # Features sheet — drop Photo column
        df_features_dl = df_features.drop(columns=['Photo'], errors='ignore')
        df_pp_dl       = df_pp.drop(columns=['Photo'], errors='ignore')

        ws1 = wb.active
        ws1.title = 'Features'
        headers = list(df_features_dl.columns)
        ws1.append(headers)
        for cell in ws1[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='051b63')
            cell.font = Font(bold=True, color='FFFFFF')
        for _, row in df_features_dl.iterrows():
            ws1.append([str(v) if v is not None else '' for v in row])

        # Passing Places sheet — drop Photo column
        ws2 = wb.create_sheet('Passing Places')
        headers2 = list(df_pp_dl.columns)
        ws2.append(headers2)
        for cell in ws2[1]:
            cell.fill = PatternFill('solid', fgColor='051b63')
            cell.font = Font(bold=True, color='FFFFFF')
        for _, row in df_pp_dl.iterrows():
            ws2.append([str(v) if v is not None else '' for v in row])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{route_id}_survey_data.xlsx"'
        return resp

    feature_types = sorted(df_features['Feature Type'].dropna().unique())
    ch_min_val = float(df_features['Chainage (m)'].min())
    ch_max_val = float(df_features['Chainage (m)'].max())

    context = {
        'routes':        routes,
        'first_route':   list(routes.keys())[0] if routes else None,
        'route_id':      route_id,
        'info':          info,
        'items':         items,
        'feature_types': feature_types,
        'ch_min_val':    ch_min_val,
        'ch_max_val':    ch_max_val,
        'q_id':          q_id,
        'q_type':        q_type,
        'ch_min':        ch_min,
        'ch_max':        ch_max,
        'photos_only':   photos_only,
        'total':         len(items),
        'with_photos':   sum(1 for i in items if i['has_photo']),
    }
    return render(request, 'survey/photo_library.html', context)

def route_summary(request, route_id):
    import json
    import math
    from collections import Counter
    import pandas as pd

    routes = scan_routes()
    if route_id not in routes:
        from django.http import Http404
        raise Http404(f"Route {route_id} not found")

    info = routes[route_id]
    df_features, df_pp, summary = load_route_data(info['xlsx'])
    first_route = list(routes.keys())[0] if routes else None

    def safe_float(val):
        try:
            v = float(val)
            return None if math.isnan(v) else v
        except:
            return None

    # ── GPS Accuracy ──────────────────────────────────────────
    acc_vals = [safe_float(v) for v in df_features['GPS Accuracy (m)'] if safe_float(v) is not None]
    pp_acc   = [safe_float(v) for v in df_pp['GPS Accuracy (m)']       if safe_float(v) is not None]
    all_acc  = acc_vals + pp_acc

    avg_acc  = round(sum(all_acc) / len(all_acc), 2) if all_acc else None
    best_acc = round(min(all_acc), 2) if all_acc else None
    worst_acc= round(max(all_acc), 2) if all_acc else None

    bands = {'< 5m': 0, '5–10m': 0, '10–15m': 0, '> 15m': 0}
    for v in all_acc:
        if   v < 5:  bands['< 5m']   += 1
        elif v < 10: bands['5–10m']   += 1
        elif v < 15: bands['10–15m']  += 1
        else:        bands['> 15m']   += 1

    # ── Entry Method — handled below with defaults ──────────

    # ── Survey dates & productivity ───────────────────────────
    df_features['_dt'] = pd.to_datetime(df_features['Captured At'], errors='coerce')
    df_pp['_dt']       = pd.to_datetime(df_pp['Captured At'],       errors='coerce')

    all_dt = pd.concat([df_features['_dt'], df_pp['_dt']]).dropna()

    date_range_start = all_dt.min().strftime('%d %b %Y') if not all_dt.empty else '—'
    date_range_end   = all_dt.max().strftime('%d %b %Y') if not all_dt.empty else '—'
    survey_days      = all_dt.dt.date.nunique() if not all_dt.empty else 0

    # Points per day
    per_day = (
        all_dt.dt.date.value_counts()
        .sort_index()
        .reset_index()
    )
    per_day.columns = ['date', 'count']
    per_day['date'] = per_day['date'].astype(str)
    per_day_json = per_day.to_dict('records')

    # Captures by hour
    per_hour = all_dt.dt.hour.value_counts().sort_index()
    per_hour_json = [{'hour': int(h), 'count': int(c)} for h, c in per_hour.items()]

    # Peak hour
    peak_hour = int(per_hour.idxmax()) if not per_hour.empty else None
    peak_hour_str = f"{peak_hour:02d}:00–{peak_hour+1:02d}:00" if peak_hour is not None else '—'

    # Surveyors
    surveyors = {}
    if 'Captured By' in df_features.columns:
        surveyors = df_features['Captured By'].value_counts().to_dict()

    # ── Coverage & Photos ─────────────────────────────────────
    n_features = len(df_features)
    n_pp       = len(df_pp)

    def count_photos(df):
        total = 0
        with_photo = 0
        for v in df['Photo'].fillna(''):
            photos = [p.strip() for p in str(v).split(',') if p.strip()]
            if photos:
                with_photo += 1
                total += len(photos)
        return with_photo, total

    feat_with_photo, feat_total_photos = count_photos(df_features)
    pp_with_photo,   pp_total_photos   = count_photos(df_pp)

    total_photos   = feat_total_photos + pp_total_photos
    total_items    = n_features + n_pp
    items_w_photos = feat_with_photo + pp_with_photo
    avg_photos     = round(total_photos / items_w_photos, 1) if items_w_photos > 0 else 0

    feat_photo_pct = round(feat_with_photo / n_features * 100) if n_features > 0 else 0
    pp_photo_pct   = round(pp_with_photo   / n_pp       * 100) if n_pp       > 0 else 0

    # ── Feature breakdown ─────────────────────────────────────
    # Feature type counts
    tc = df_features['Feature Type'].value_counts().reset_index()
    tc.columns = ['type', 'count']
    type_counts_list = [{'type': str(r['type']), 'count': int(r['count'])}
                        for _, r in tc.iterrows()]
    type_max = max((x['count'] for x in type_counts_list), default=1)

    # Condition counts
    cond_list = [{'condition': k, 'count': int(v)}
                 for k, v in df_features['Condition'].value_counts().items()]

    # Route length — DXF if available, chainage range fallback
    _length = get_route_length_m(info, df_features)
    route_length_m = f"{_length:,}" if _length else '—' 

    # Est. on-site time — round first capture DOWN and last UP to whole hours per day
    import math
    survey_hours = '—'
    if not all_dt.empty:
        total_hours = 0
        for day, group in all_dt.groupby(all_dt.dt.date):
            first_hour = math.floor(group.min().hour + group.min().minute / 60)
            last_hour  = math.ceil(group.max().hour + group.max().minute / 60)
            total_hours += last_hour - first_hour
        survey_hours = str(total_hours)

    # Entry method — ensure Manual shows 0 if not present
    if 'Entry Method' in df_features.columns:
        all_methods = df_features['Entry Method'].value_counts().to_dict()
        for m in ['GPS', 'Manual']:
            if m not in all_methods:
                all_methods[m] = 0
        entry_counts = all_methods
    else:
        entry_counts = {'GPS': len(df_features), 'Manual': 0}

    # Per day max for bar scaling
    per_day_max = max((d['count'] for d in per_day_json), default=1)

    # Per hour by day — combine features + passing places
    df_pp['_dt'] = pd.to_datetime(df_pp['Captured At'], errors='coerce')
    all_dt_combined = pd.concat([
        df_features['_dt'].dropna(),
        df_pp['_dt'].dropna()
    ])
    per_hour_by_day = {}
    for dt_val, group in all_dt_combined.groupby(all_dt_combined.dt.date):
        hour_counts = group.dt.hour.value_counts().sort_index()
        per_hour_by_day[str(dt_val)] = [
            {'hour': h, 'count': int(hour_counts.get(h, 0))}
            for h in range(6, 20)   # show 06:00 to 19:00
        ]
    per_hour_max_all = max(
        (h['count'] for day_data in per_hour_by_day.values() for h in day_data),
        default=1
    )

    # Coverage & productivity metrics
    total_points   = n_features + n_pp
    length_m       = float(_length) if _length else None
    s_hours        = float(survey_hours) if survey_hours not in ('—', '0') else None

    # Use surveyed chainage range for productivity metrics (not full DXF length)
    surveyed_length = float(df_features['Chainage (m)'].max()) - float(df_features['Chainage (m)'].min())

    pts_per_100m    = round(total_points / surveyed_length * 100, 1) if surveyed_length > 0 else None
    pts_per_hour    = round(total_points / s_hours, 1)               if s_hours and s_hours > 0 else None
    metres_per_hour = round(surveyed_length / s_hours)               if s_hours and s_hours > 0 else None

    # Metres covered per day
    df_features['_dt'] = pd.to_datetime(df_features['Captured At'], errors='coerce')
    metres_per_day = {}
    for date, group in df_features.dropna(subset=['_dt', 'Chainage (m)']).groupby(df_features['_dt'].dt.date):
        ch_range = float(group['Chainage (m)'].max()) - float(group['Chainage (m)'].min())
        metres_per_day[str(date)] = round(ch_range)

    # ── Passing places spacing helper ────────────────────────────
    def _build_pp_spacing(df_pp):
        import math
        rows = []
        df_sorted = df_pp.sort_values('Mid Chainage (m)').reset_index(drop=True)
        for i, row in df_sorted.iterrows():
            ch = row.get('Mid Chainage (m)')
            try:
                ch = float(ch) if ch and not math.isnan(float(ch)) else None
            except:
                ch = None
            spacing = None
            if i > 0 and ch is not None:
                prev_ch = rows[-1]['chainage'] if rows else None
                if prev_ch is not None:
                    spacing = round(ch - prev_ch)
            rows.append({
                'id':       str(row.get('PP ID', '')),
                'chainage': round(ch) if ch else None,
                'status':   str(row.get('Status', '')),
                'side':     str(row.get('Side', '')),
                'spacing':  spacing,
            })
        return rows

    # ── Survey progress from DXF ─────────────────────────────────
    dxf_length    = float(_length) if _length else None
    first_chainage = float(df_features['Chainage (m)'].min())
    last_chainage  = float(df_features['Chainage (m)'].max())
    surveyed_length_m = last_chainage - first_chainage

    progress_pct     = round(surveyed_length_m / dxf_length * 100, 1) if dxf_length else None
    start_offset_m   = round(first_chainage)
    end_offset_m     = round(dxf_length - last_chainage) if dxf_length else None

    # ── Gap detection ────────────────────────────────────────────
    from django.conf import settings as django_settings

    # Get threshold from request (form submit) or settings default
    gap_threshold = int(request.GET.get('gap_threshold',
                        getattr(django_settings, 'SURVEY_GAP_THRESHOLD', 8)))

    # Merge all captured points by chainage (features + passing places)
    all_points = []
    for _, row in df_features.iterrows():
        ch = row.get('Chainage (m)')
        if ch and not math.isnan(float(ch)):
            all_points.append({
                'id':       str(row['ID']),
                'kind':     'Feature',
                'type':     str(row.get('Feature Type', '')),
                'chainage': float(ch),
            })
    for _, row in df_pp.iterrows():
        ch = row.get('Mid Chainage (m)')
        if ch and not math.isnan(float(ch)):
            all_points.append({
                'id':       str(row['PP ID']),
                'kind':     'Passing Place',
                'type':     'Passing Place',
                'chainage': float(ch),
            })

    all_points.sort(key=lambda x: x['chainage'])

    # Calculate gaps between consecutive points
    gaps = []
    spacings = []
    for i in range(1, len(all_points)):
        spacing = all_points[i]['chainage'] - all_points[i-1]['chainage']
        spacings.append(spacing)
        gaps.append({
            'from_id':      all_points[i-1]['id'],
            'from_type':    all_points[i-1]['kind'],
            'to_id':        all_points[i]['id'],
            'to_type':      all_points[i]['kind'],
            'from_chainage': all_points[i-1]['chainage'],
            'to_chainage':   all_points[i]['chainage'],
            'spacing':       round(spacing, 1),
        })

    avg_spacing = round(sum(spacings) / len(spacings), 1) if spacings else 0
    threshold_m = round(avg_spacing * gap_threshold, 1)

    # Build lat/lon lookup from all points for gap map
    ch_to_latlon = {}
    for _, row in df_features.iterrows():
        ch = row.get('Chainage (m)')
        if ch and not math.isnan(float(ch)):
            ch_to_latlon[float(ch)] = {
                'lat': row.get('Latitude'), 'lon': row.get('Longitude')
            }
    for _, row in df_pp.iterrows():
        ch = row.get('Mid Chainage (m)')
        if ch and not math.isnan(float(ch)):
            ch_to_latlon[float(ch)] = {
                'lat': row.get('Mid Latitude'), 'lon': row.get('Mid Longitude')
            }

    flagged_gaps = []
    for g in gaps:
        if avg_spacing > 0 and g['spacing'] >= threshold_m:
            from_ll = ch_to_latlon.get(g['from_chainage'], {})
            to_ll   = ch_to_latlon.get(g['to_chainage'],   {})
            flagged_gaps.append({
                **g,
                'multiple':  round(g['spacing'] / avg_spacing, 1),
                'from_lat':  from_ll.get('lat'),
                'from_lon':  from_ll.get('lon'),
                'to_lat':    to_ll.get('lat'),
                'to_lon':    to_ll.get('lon'),
            })
    flagged_gaps.sort(key=lambda x: x['spacing'], reverse=True)

    # Generate coverage map
    from .services.gap_map import get_coverage_map
    dxf_coords    = get_alignment_coords(info, df_features)
    survey_points = sorted(
        [
            {'lat': ch_to_latlon[p['chainage']]['lat'],
             'lon': ch_to_latlon[p['chainage']]['lon'],
             'chainage': p['chainage']}
            for p in all_points
            if p['chainage'] in ch_to_latlon
            and ch_to_latlon[p['chainage']].get('lat')
            and ch_to_latlon[p['chainage']].get('lon')
        ],
        key=lambda x: x['chainage']
    )

    # Regenerate map if threshold changed via form
    if request.GET.get('gap_threshold'):
        from .services.gap_map import clear_coverage_map
        clear_coverage_map(route_id)
    pp_for_map = [
        {'lat': float(r['Mid Latitude']), 'lon': float(r['Mid Longitude'])}
        for _, r in df_pp.iterrows()
        if r.get('Mid Latitude') and r.get('Mid Longitude')
    ]
    coverage_map_url = get_coverage_map(route_id, dxf_coords, survey_points,
                                        flagged_gaps, passing_places=pp_for_map)

    context = {
        'routes':           routes,
        'route_id':         route_id,
        'first_route':      first_route,
        'info':             info,
        # Quick header stats
        'route_length_m':   route_length_m,
        'pts_per_100m':     pts_per_100m,
        # Progress
        'dxf_length':       round(dxf_length) if dxf_length else None,
        'surveyed_length_m': round(surveyed_length_m),
        'progress_pct':     progress_pct,
        'start_offset_m':   start_offset_m,
        'end_offset_m':     end_offset_m,
        # Gap detection
        'avg_spacing':      avg_spacing,
        'threshold_m':      threshold_m,
        'gap_threshold':    gap_threshold,
        'flagged_gaps':     flagged_gaps,
        'coverage_map_url': coverage_map_url,
        # Passing places spacing table
        'pp_spacing':       _build_pp_spacing(df_pp),
        'pp_avg_spacing':   round(sum(p['spacing'] for p in _build_pp_spacing(df_pp) if p['spacing'])
                            / max(1, sum(1 for p in _build_pp_spacing(df_pp) if p['spacing']))),
        'total_points':     len(all_points),
        'metres_per_hour':  metres_per_hour,
        'metres_per_day':     metres_per_day,
        'metres_per_day_max': max(metres_per_day.values()) if metres_per_day else 1,
        'pts_per_hour':     pts_per_hour,
        'survey_hours':     survey_hours,
        # GPS
        'avg_acc':          avg_acc,
        'best_acc':         best_acc,
        'worst_acc':        worst_acc,
        'bands':            bands,
        'entry_counts':     entry_counts,
        # Time
        'date_range_start': date_range_start,
        'date_range_end':   date_range_end,
        'survey_days':      survey_days,
        'per_day_list':       per_day_json,
        'per_day_max':        per_day_max,
        'per_hour_by_day':    per_hour_by_day,
        'per_hour_max_all':   per_hour_max_all,
        'peak_hour_str':    peak_hour_str,
        'surveyors':        surveyors,
        # Coverage
        'n_features':       n_features,
        'n_pp':             n_pp,
        'feat_with_photo':  feat_with_photo,
        'feat_photo_pct':   feat_photo_pct,
        'pp_with_photo':    pp_with_photo,
        'pp_photo_pct':     pp_photo_pct,
        'total_photos':     total_photos,
        'avg_photos':       avg_photos,
        # Feature breakdown
        'type_counts_list': type_counts_list,
        'type_max':         type_max,
        'cond_list':        cond_list,
        'n_good':           int((df_features['Condition'] == 'GOOD').sum()),
        'n_fair':           int((df_features['Condition'] == 'FAIR').sum()),
        'n_poor':           int((df_features['Condition'] == 'POOR').sum()),
    }
    return render(request, 'survey/summary.html', context)
